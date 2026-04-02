"""Generate ablation results Excel from summary JSONs."""
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent

# Load data
old = json.loads((BASE / "ablation/ablation_summary.json").read_text())
new_c5 = json.loads((BASE / "ablation_c5_rerun/ablation_summary.json").read_text())

rows = []
for mode in ["full", "ablate_c1", "ablate_c2", "ablate_c3", "ablate_c4"]:
    s = old["summaries"][mode]
    rows.append(("old", mode, s))

# old c5
rows.append(("old", "ablate_c5 (旧-纯规则)", old["summaries"]["ablate_c5"]))
# new c5
rows.append(("new", "ablate_c5 (修复-LLM worker)", new_c5["summaries"]["ablate_c5"]))

COMPONENT = {
    "full": "全部启用",
    "ablate_c1": "禁用 C1 JSON Patch",
    "ablate_c2": "禁用 C2 Schema验证",
    "ablate_c3": "禁用 C3 熔断器",
    "ablate_c4": "禁用 C4 上下文裁剪",
    "ablate_c5 (旧-纯规则)": "禁用 C5 LLM Architect（旧）",
    "ablate_c5 (修复-LLM worker)": "禁用 C5 LLM Architect（修复）",
}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "消融实验结果"

headers = [
    "模式", "禁用组件",
    "成功率", "平均步数",
    "平均总Tokens", "Worker Tokens", "Architect Tokens",
    "熔断触发率", "停滞率", "Fallback率", "PatchError率",
    "成功局数", "超步局数", "停滞局数",
]

# Header style
header_fill = PatternFill("solid", fgColor="2F5496")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

ws.row_dimensions[1].height = 36

# Row fills
fills = {
    "full":                          PatternFill("solid", fgColor="E2EFDA"),
    "ablate_c1":                     PatternFill("solid", fgColor="FCE4D6"),
    "ablate_c2":                     PatternFill("solid", fgColor="EBF3FB"),
    "ablate_c3":                     PatternFill("solid", fgColor="EBF3FB"),
    "ablate_c4":                     PatternFill("solid", fgColor="FFF2CC"),
    "ablate_c5 (旧-纯规则)":          PatternFill("solid", fgColor="F2F2F2"),
    "ablate_c5 (修复-LLM worker)":   PatternFill("solid", fgColor="E2EFDA"),
}

def pct(v): return f"{v*100:.1f}%"
def num(v): return round(v, 1)

for row_idx, (_, mode, s) in enumerate(rows, 2):
    stop = s.get("stop_reason_breakdown", {})
    data = [
        mode,
        COMPONENT.get(mode, ""),
        pct(s["success_rate"]),
        num(s["mean_steps"]),
        num(s["mean_total_tokens"]),
        num(s["mean_worker_input_tokens"] + s["mean_worker_output_tokens"]),
        num(s["mean_architect_total_tokens"]),
        pct(s["circuit_breaker_trigger_rate"]),
        pct(s["mean_stagnation_rate"]),
        pct(s["mean_fallback_rate"]),
        pct(s["mean_patch_error_rate"]),
        stop.get("environment_done", 0),
        stop.get("max_steps_reached", 0),
        stop.get("stagnation_detected", 0),
    ]
    fill = fills.get(mode, PatternFill("solid", fgColor="FFFFFF"))
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

# Column widths
col_widths = [28, 26, 10, 10, 14, 14, 16, 12, 10, 10, 12, 10, 10, 10]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"

out = BASE / "ablation_results.xlsx"
wb.save(out)
print(f"Saved: {out}")

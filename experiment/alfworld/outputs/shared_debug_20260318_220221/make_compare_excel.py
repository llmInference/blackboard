"""Generate three-system comparison Excel for ALFWorld debug run."""
import json
import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent

# ── data paths ────────────────────────────────────────────────────────────────
PATHS = {
    "blackboard": BASE / "ablation/full/full_episodes.jsonl",
    "langgraph":  BASE / "system_compare/langgraph/langgraph_episodes.jsonl",
    "autogen":    BASE / "system_compare/autogen/autogen_episodes.jsonl",
}
SUMMARY_PATHS = {
    "blackboard": BASE / "ablation/full/full_summary.json",
    "langgraph":  BASE / "system_compare/langgraph/langgraph_summary.json",
    "autogen":    BASE / "system_compare/autogen/autogen_summary.json",
}

# ── helpers ───────────────────────────────────────────────────────────────────
def load_episodes(path):
    eps = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                eps.append(json.loads(line))
    return eps

def task_key(ep):
    gf = ep.get("gamefile") or ep.get("task_id") or ""
    # use the trial directory name as a short key
    parts = Path(gf).parts
    for i, p in enumerate(parts):
        if p.startswith("trial_"):
            return "/".join(parts[max(0, i-1):i+1])
    return Path(gf).stem

def fmt_pct(v):
    return f"{v*100:.1f}%"

def fmt_f(v, d=1):
    return round(float(v), d)

# ── styles ────────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=11)
SYS_FILLS  = {
    "blackboard": PatternFill("solid", fgColor="D6E4F0"),
    "langgraph":  PatternFill("solid", fgColor="D5E8D4"),
    "autogen":    PatternFill("solid", fgColor="FFF2CC"),
}
SYS_HDR_FILLS = {
    "blackboard": PatternFill("solid", fgColor="2E75B6"),
    "langgraph":  PatternFill("solid", fgColor="548235"),
    "autogen":    PatternFill("solid", fgColor="BF8F00"),
}
SUCCESS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL    = PatternFill("solid", fgColor="FFCCCC")
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hdr_cell(ws, row, col, value, fill=None, font=None, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    if fill: c.fill = fill
    if font: c.font = font
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = BORDER
    return c

def data_cell(ws, row, col, value, fill=None, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    if fill: c.fill = fill
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = BORDER
    return c

# ── Sheet 1: Summary ──────────────────────────────────────────────────────────
def build_summary_sheet(wb):
    ws = wb.create_sheet("Summary")
    summaries = {k: json.loads(Path(v).read_text()) for k, v in SUMMARY_PATHS.items()}

    metrics = [
        ("n_episodes",              "Episodes",           False),
        ("success_rate",            "Success Rate",       True),
        ("mean_steps",              "Avg Steps",          False),
        ("mean_total_tokens",       "Avg Total Tokens",   False),
        ("mean_worker_input_tokens","Avg Worker In Tok",  False),
        ("mean_worker_output_tokens","Avg Worker Out Tok",False),
        ("mean_architect_input_tokens","Avg Arch In Tok", False),
        ("mean_architect_output_tokens","Avg Arch Out Tok",False),
        ("mean_fallback_rate",      "Fallback Rate",      True),
        ("mean_repeat_action_rate", "Repeat Action Rate", True),
        ("mean_stagnation_rate",    "Stagnation Rate",    True),
        ("circuit_breaker_trigger_rate","Circuit Breaker Rate",True),
    ]

    systems = ["blackboard", "langgraph", "autogen"]

    # header row
    hdr_cell(ws, 1, 1, "Metric", HDR_FILL, HDR_FONT)
    for ci, sys in enumerate(systems, 2):
        hdr_cell(ws, 1, ci, sys.capitalize(), SYS_HDR_FILLS[sys],
                 Font(bold=True, color="FFFFFF", size=11))

    for ri, (key, label, is_pct) in enumerate(metrics, 2):
        data_cell(ws, ri, 1, label, align="left")
        vals = [summaries[s].get(key, 0) for s in systems]
        for ci, (sys, v) in enumerate(zip(systems, vals), 2):
            txt = fmt_pct(v) if is_pct else fmt_f(v, 0 if key == "n_episodes" else 1)
            fill = SYS_FILLS[sys]
            # highlight best success rate
            if key == "success_rate" and v == max(vals):
                fill = SUCCESS_FILL
            data_cell(ws, ri, ci, txt, fill)

    # stop reason breakdown
    row = len(metrics) + 3
    hdr_cell(ws, row, 1, "Stop Reason Breakdown", HDR_FILL, HDR_FONT)
    for ci, sys in enumerate(systems, 2):
        hdr_cell(ws, row, ci, sys.capitalize(), SYS_HDR_FILLS[sys],
                 Font(bold=True, color="FFFFFF", size=11))
    row += 1
    all_reasons = set()
    for s in systems:
        all_reasons |= set(summaries[s].get("stop_reason_breakdown", {}).keys())
    for reason in sorted(all_reasons):
        data_cell(ws, row, 1, reason, align="left")
        for ci, sys in enumerate(systems, 2):
            v = summaries[sys].get("stop_reason_breakdown", {}).get(reason, 0)
            data_cell(ws, row, ci, v, SYS_FILLS[sys])
        row += 1

    ws.column_dimensions["A"].width = 26
    for ci in range(2, 5):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.freeze_panes = "B2"

# ── Sheet 2: Per-Episode Detail ───────────────────────────────────────────────
def build_episode_sheet(wb):
    ws = wb.create_sheet("Episodes")
    systems = ["blackboard", "langgraph", "autogen"]
    all_eps = {s: load_episodes(PATHS[s]) for s in systems}

    # build task key → episode map per system
    ep_map = {s: {task_key(e): e for e in all_eps[s]} for s in systems}
    all_keys = []
    seen = set()
    for s in systems:
        for e in all_eps[s]:
            k = task_key(e)
            if k not in seen:
                all_keys.append(k)
                seen.add(k)

    # columns: task | [success, steps, tokens, status] x3
    col_groups = [
        ("Success", "success"),
        ("Steps",   "steps"),
        ("Tokens",  "total_tokens"),
        ("Status",  "final_status"),
    ]
    n_cols = len(col_groups)

    # row 1: system headers (merged)
    hdr_cell(ws, 1, 1, "Task", HDR_FILL, HDR_FONT, align="left")
    for si, sys in enumerate(systems):
        start_col = 2 + si * n_cols
        hdr_cell(ws, 1, start_col, sys.capitalize(),
                 SYS_HDR_FILLS[sys], Font(bold=True, color="FFFFFF", size=11))
        ws.merge_cells(start_row=1, start_column=start_col,
                       end_row=1, end_column=start_col + n_cols - 1)

    # row 2: column sub-headers
    hdr_cell(ws, 2, 1, "Task", HDR_FILL, HDR_FONT, align="left")
    for si, sys in enumerate(systems):
        for ci, (label, _) in enumerate(col_groups):
            col = 2 + si * n_cols + ci
            hdr_cell(ws, 2, col, label, SYS_HDR_FILLS[sys],
                     Font(bold=True, color="FFFFFF", size=10))

    # data rows
    for ri, key in enumerate(all_keys, 3):
        data_cell(ws, ri, 1, key, align="left")
        for si, sys in enumerate(systems):
            ep = ep_map[sys].get(key)
            for ci, (_, field) in enumerate(col_groups):
                col = 2 + si * n_cols + ci
                if ep is None:
                    data_cell(ws, ri, col, "—")
                    continue
                v = ep.get(field)
                if field == "success":
                    txt = "✓" if v else "✗"
                    fill = SUCCESS_FILL if v else FAIL_FILL
                elif field == "total_tokens":
                    txt = int(v) if v is not None else 0
                    fill = SYS_FILLS[sys]
                else:
                    txt = v
                    fill = SYS_FILLS[sys]
                data_cell(ws, ri, col, txt, fill)

    ws.column_dimensions["A"].width = 55
    for si in range(len(systems)):
        for ci in range(n_cols):
            col = 2 + si * n_cols + ci
            ws.column_dimensions[get_column_letter(col)].width = 12
    ws.freeze_panes = "B3"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20

# ── Sheet 3: Task-type breakdown ──────────────────────────────────────────────
def build_tasktype_sheet(wb):
    ws = wb.create_sheet("By Task Type")
    systems = ["blackboard", "langgraph", "autogen"]
    all_eps = {s: load_episodes(PATHS[s]) for s in systems}

    def task_type(ep):
        gf = ep.get("gamefile") or ""
        for part in Path(gf).parts:
            if "-" in part and not part.startswith("trial_"):
                return part.split("-")[0]
        return "unknown"

    # collect types
    type_eps = {}
    for s in systems:
        for e in all_eps[s]:
            t = task_type(e)
            type_eps.setdefault(t, {s: [] for s in systems})
            type_eps[t][s].append(e)

    hdr_cell(ws, 1, 1, "Task Type", HDR_FILL, HDR_FONT, align="left")
    for ci, sys in enumerate(systems, 2):
        hdr_cell(ws, 1, ci, f"{sys.capitalize()}\nSuccess Rate",
                 SYS_HDR_FILLS[sys], Font(bold=True, color="FFFFFF", size=11))
    hdr_cell(ws, 1, 5, "N (per system)", HDR_FILL, HDR_FONT)

    for ri, (ttype, sys_eps) in enumerate(sorted(type_eps.items()), 2):
        data_cell(ws, ri, 1, ttype, align="left")
        for ci, sys in enumerate(systems, 2):
            eps = sys_eps.get(sys, [])
            if not eps:
                data_cell(ws, ri, ci, "—")
            else:
                rate = sum(1 for e in eps if e.get("success")) / len(eps)
                fill = SUCCESS_FILL if rate >= 0.5 else FAIL_FILL
                data_cell(ws, ri, ci, fmt_pct(rate), fill)
        n = max(len(sys_eps.get(s, [])) for s in systems)
        data_cell(ws, ri, 5, n)

    ws.column_dimensions["A"].width = 35
    for ci in range(2, 6):
        ws.column_dimensions[get_column_letter(ci)].width = 18
    ws.freeze_panes = "B2"

# ── main ──────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_summary_sheet(wb)
    build_episode_sheet(wb)
    build_tasktype_sheet(wb)

    out = BASE / "system_compare_3way.xlsx"
    wb.save(out)
    print(f"Saved: {out}")

if __name__ == "__main__":
    main()
